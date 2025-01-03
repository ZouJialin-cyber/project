"""
    kanban:
    根据本地数据贴图到excel表,
    准备好需要贴图的数据, 这里会把这些数据resize到200分辨率, 然后将其贴到看板
"""

import pandas as pd
import cv2
import tifffile as tif
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
from tqdm import tqdm


# 读取Excel文件
def read_excel(file_path):
    try:
        df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"读取Excel文件时出错: {e}")
        return None


def f_ij_16_to_8(img, chunk_size=1000):
    if img.dtype == 'uint8':
        return img
    dst = np.zeros(img.shape, np.uint8)
    p_max = np.max(img)
    p_min = np.min(img)
    scale = 256.0 / (p_max - p_min + 1)
    for idx in range(img.shape[0] // chunk_size + 1):
        sl = slice(idx * chunk_size, (idx + 1) * chunk_size)
        win_img = np.copy(img[sl])
        win_img = np.int16(win_img)
        win_img = (win_img & 0xffff)
        win_img = win_img - p_min
        win_img[win_img < 0] = 0
        win_img = win_img * scale + 0.5
        win_img[win_img > 255] = 255
        dst[sl] = np.array(win_img).astype(np.uint8)
    return dst


# 降采样图像
def downsample_image(image_path, target_width=200, target_height=200):
    try:
        # image = tif.imread(image_path)
        image = cv2.imread(image_path, -1)
        image = f_ij_16_to_8(image)
        if image is None:
            print(f"无法读取图像 {image_path}")
            return None
        downsampled_image = cv2.resize(image, (target_width, target_height), interpolation=cv2.INTER_NEAREST)
        return downsampled_image
    except Exception as e:
        print(f"降采样图像时出错: {e}")
        return None


# 将图像插入到Excel单元格中
def insert_image(sheet, start_column_index, start_row_index, image_path):
    cell_width = 40  # 设置单元格宽度，可根据实际需求调整
    cell_height = 200  # 设置单元格高度，可根据实际需求调整

    sheet.column_dimensions[get_column_letter(start_column_index)].width = cell_width
    sheet.row_dimensions[start_row_index].height = cell_height
    img = Image(image_path)
    src = AnchorMarker(col=start_column_index - 1, colOff=0, row=start_row_index - 1, rowOff=0)
    end = AnchorMarker(col=start_column_index, colOff=0, row=start_row_index, rowOff=0)

    img.anchor = TwoCellAnchor('twoCell', src, end)

    sheet.add_image(img)


# 主函数
def main():
    # Excel文件路径
    excel_file_path = r"C:\Users\zoujialin\Desktop\银标数据的入库总表.xlsx"

    # 图像文件夹路径（假设图像都在这个文件夹下，根据实际情况修改）
    image_folder_path = r"C:\Users\zoujialin\Desktop\unnormal_data"

    # 读取Excel文件
    excel_data = read_excel(excel_file_path)
    if excel_data is None:
        return

    # 加载Excel文档
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # 新的图像保存目的地址
    new_image_folder_path = r"C:\Users\zoujialin\Desktop\temporary"
    if not os.path.exists(new_image_folder_path):
        os.makedirs(new_image_folder_path)

    # 遍历Excel表格中的每一行（假设前缀在第一列）
    for index, row in tqdm(excel_data.iterrows(), total=len(excel_data), desc="处理数据"):
        prefix = row[0]  # 获取第一列的前缀

        # 在图像文件夹中查找以该前缀命名的图像文件
        for file in os.listdir(image_folder_path):
            if file.startswith(prefix):
                image_path = os.path.join(image_folder_path, file)
                break
        else:
            print(f"未找到前缀 {prefix} 对应的图像文件")
            continue

        # 降采样图像
        downsampled_image = downsample_image(image_path)
        if downsampled_image is not None:
            # 保存降采样后的图像到新的目的地址，图像名按照前缀命名
            new_image_path = os.path.join(new_image_folder_path, f'{prefix}.png')
            cv2.imwrite(new_image_path, downsampled_image)

    # 再次遍历Excel表格，将处理后的图像插入到Excel单元格中
    for index, row in excel_data.iterrows():
        prefix = row[0]
        row_index = index + 1 + 1  # excel第一行从1开始，第一行是表头不插入，因此加2
        try:
            new_image_path = os.path.join(new_image_folder_path, f'{prefix}.png')    # 这里的文件路径需要根据具体需求修改
            if os.path.exists(new_image_path):
                # 获取行索引（注意：这里的索引从0开始，与Excel中的行号对应需要加1）
                # row_index = index + 1 + 1    # excel第一行从1开始，第一行是表头不插入，因此加2
                # 获取列索引（假设G列对应索引为7）
                column_index = 7
                insert_image(sheet, column_index, row_index, new_image_path)
                print(f"已将图像 {prefix}.png 插入到单元格 G{row_index}")
            else:
                print(f"未找到处理后的图像 {prefix}.png ，无法插入到单元格 G{row_index}")
        except Exception as e:
            print(f"插入图像到单元格 G{row_index} 时出错: {e}")

    # 保存修改后的Excel文件
    modified_excel_file_path = r"C:\Users\zoujialin\Desktop\银标数据的入库总表_xiugai.xlsx"
    workbook.save(modified_excel_file_path)
    print(f"已保存修改后的Excel文件到 {modified_excel_file_path}")


if __name__ == "__main__":
    main()